import csv
import json
import os
import io
import zipfile
import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Count, Q, Avg, Sum, Max
from django.db import transaction
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt

# We might not have Anthropic installed yet or configured, handling it gracefully
try:
    from anthropic import Anthropic
except ImportError:
    Anthropic = None

from .models import Drive, Student, Company, Profile, Branch, InterviewRound, RoundStudent
from .decorators import login_required
from .forms import (
    DriveStep1Form, BranchSelectionForm, StudentForm, 
    CompanyForm, ProfileForm, CSVUploadForm, BranchForm, EditDriveForm,
    InterviewRoundForm
)

@login_required
def dashboard(request):
    """Stats cards + chart data per drive"""
    active_drives = Drive.objects.filter(status='active')
    
    drive_stats = []
    for d in active_drives:
        drive_stats.append({
            'drive': d,
            'total_students': d.students.count(),
            'placed_students': d.students.exclude(placement_status='Unplaced').count(),
            'total_companies': d.companies.count()
        })
    
    context = {
        'drive_stats': drive_stats,
    }
    return render(request, 'portal/dashboard.html', context)


@login_required
def drives_list(request):
    """All drives table"""
    drives = Drive.objects.all()
    branches = Branch.objects.all()
    return render(request, 'portal/drives.html', {'drives': drives, 'branches': branches})


@login_required
def create_drive_step1(request):
    if request.method == 'POST':
        form = DriveStep1Form(request.POST)
        if form.is_valid():
            drive = form.save(commit=False)
            try:
                drive.clean()
                drive.save()
                request.session['latest_drive_id'] = drive.drive_id
                return redirect('create_drive_step2')
            except ValidationError as e:
                messages.error(request, e.message)
    else:
        form = DriveStep1Form()
    return render(request, 'portal/create_drive_step1.html', {'form': form})


@login_required
def create_drive_step2(request):
    drive_id = request.session.get('latest_drive_id')
    if drive_id:
        latest_drive = Drive.objects.filter(pk=drive_id).first()
    else:
        latest_drive = Drive.objects.order_by('-drive_year').first()
        
    if not latest_drive:
        return redirect('create_drive_step1')
        
    if request.method == 'POST':
        form = BranchSelectionForm(request.POST)
        if form.is_valid():
            branches = form.cleaned_data['branches']
            latest_drive.branches.set(branches)
            messages.success(request, 'Drive created successfully!')
            return redirect('drives')
    else:
        form = BranchSelectionForm()
    return render(request, 'portal/create_drive_step2.html', {'form': form, 'drive': latest_drive})


@login_required
def toggle_drive_status(request, drive_id):
    if request.method == 'POST':
        drive = get_object_or_404(Drive, pk=drive_id)
        if drive.status == 'active':
            drive.status = 'inactive'
            drive.save()
            messages.success(request, f'Drive {drive.drive_name} deactivated (archived).')
        else:
            try:
                drive.status = 'active'
                drive.clean()
                drive.save()
                messages.success(request, f'Drive {drive.drive_name} activated.')
            except ValidationError as e:
                messages.error(request, e.message)
    return redirect('drives')


@login_required
def edit_drive(request, drive_id):
    drive = get_object_or_404(Drive, pk=drive_id)
    if request.method == 'POST':
        form = EditDriveForm(request.POST, instance=drive)
        if form.is_valid():
            try:
                f = form.save(commit=False)
                f.clean()
                f.save()
                form.save_m2m()
                messages.success(request, 'Drive updated successfully.')
            except ValidationError as e:
                messages.error(request, e.message)
        else:
            for error in form.errors.values():
                messages.error(request, str(error))
        return redirect('drives')
    # Can also render a dedicated edit page or modal

@login_required
def delete_drive(request, drive_id):
    if request.method == 'POST':
        drive = get_object_or_404(Drive, pk=drive_id)
        drive_name = drive.drive_name
        drive.delete()
        messages.success(request, f'Drive {drive_name} deleted completely.')
    return redirect('drives')

@login_required
def export_drive_data(request, drive_id):
    drive = get_object_or_404(Drive, pk=drive_id)
    
    # 1. Create Students CSV
    student_output = io.StringIO()
    student_writer = csv.writer(student_output)
    student_writer.writerow(['ID', 'Name', 'Branch', 'CPI', 'Placement Status', 'Company', 'Profile Name', 'CTC (LPA)', 'Stipend', 'Offer Letter'])
    
    students = Student.all_objects.filter(drive=drive)
    for std in students:
        student_writer.writerow([
            std.student_id, std.std_name, std.branch.branch_name, 
            std.cpi, std.placement_status, 
            std.company.cmp_name if std.company else '', 
            std.profile.profile_name if std.profile else '',
            std.profile.ctc if std.profile else '',
            std.profile.stipend if std.profile else '',
            std.offer_letter or ''
        ])
    
    # 2. Create Companies CSV
    company_output = io.StringIO()
    company_writer = csv.writer(company_output)
    company_writer.writerow(['Company Name', 'Profile Name', 'CTC (LPA)', 'Stipend', 'Eligible Branches'])
    
    companies = Company.all_objects.filter(drive=drive).prefetch_related('profiles__branches')
    for cmp in companies:
        if cmp.profiles.exists():
            for profile in cmp.profiles.all():
                branches_str = ", ".join([b.branch_id for b in profile.branches.all()])
                company_writer.writerow([
                    cmp.cmp_name, profile.profile_name, profile.ctc, profile.stipend or '', branches_str
                ])
        else:
            company_writer.writerow([cmp.cmp_name, '', '', '', ''])

    # 3. Create ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        zip_file.writestr('students.csv', student_output.getvalue())
        zip_file.writestr('companies.csv', company_output.getvalue())

    # 4. Return HttpResponse
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{drive.drive_id}_export.zip"'
    return response


@login_required
def students_list(request):
    students = Student.objects.all()
    
    # Apply filters if any
    drive_id = request.GET.get('drive')
    branch_ids = request.GET.getlist('branch')
    placement_status = request.GET.get('placement_status')
    search_query = request.GET.get('search')
    cpi_min = request.GET.get('cpi_min')
    cpi_max = request.GET.get('cpi_max')
    profile_id = request.GET.get('profile')
    
    if profile_id:
        students = students.filter(profile_id=profile_id)
    if drive_id:
        students = students.filter(drive_id=drive_id)
    if branch_ids:
        clean_branches = [b for b in branch_ids if b]
        if clean_branches:
            students = students.filter(branch_id__in=clean_branches)
    if placement_status:
        students = students.filter(placement_status=placement_status)
    if search_query:
        # Search by student ID or Name
        from django.db.models import Q
        students = students.filter(
            Q(student_id__icontains=search_query) | 
            Q(std_name__icontains=search_query)
        )
    if cpi_min:
        try:
            students = students.filter(cpi__gte=float(cpi_min))
        except ValueError:
            pass
    if cpi_max:
        try:
            students = students.filter(cpi__lte=float(cpi_max))
        except ValueError:
            pass
            
    # Sort by Student ID
    students = students.order_by('student_id')
        
    paginator = Paginator(students, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
        
    context = {
        'students': page_obj,
        'drives': Drive.objects.filter(status='active'),
        'branches': Branch.objects.all(),
        'companies': Company.objects.all(),
        'profiles': Profile.objects.all(),
        'student_form': StudentForm(),
        'csv_form': CSVUploadForm(),
        'branch_form': BranchForm(),
        'selected_branches': branch_ids,
        'filter_string': query_params.urlencode()
    }
    return render(request, 'portal/students.html', context)


@login_required
def add_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student added successfully.')
        else:
            for error in form.errors.values():
                messages.error(request, str(error))
    return redirect('students')


@login_required
def add_branch(request):
    if request.method == 'POST':
        form = BranchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Branch added successfully.')
        else:
            for error in form.errors.values():
                messages.error(request, str(error))
    return redirect(request.META.get('HTTP_REFERER', 'branches'))

@login_required
def branches_list(request):
    """View all branches."""
    branches = Branch.objects.all()
    branch_form = BranchForm()
    return render(request, 'portal/branches.html', {'branches': branches, 'branch_form': branch_form})

@login_required
def edit_branch(request, branch_id):
    """Edit branch details."""
    branch = get_object_or_404(Branch, pk=branch_id)
    if request.method == 'POST':
        form = BranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, 'Branch updated successfully.')
        else:
            for error in form.errors.values():
                messages.error(request, str(error))
    return redirect('branches')

@login_required
def delete_branch(request, branch_id):
    """Delete a branch if no students are tied to it."""
    branch = get_object_or_404(Branch, pk=branch_id)
    if request.method == 'POST':
        try:
            branch.delete()
            messages.success(request, 'Branch deleted successfully.')
        except Exception as e:
            messages.error(request, 'Cannot delete this branch. Make sure no students or drives are actively using it.')
    return redirect('branches')


@login_required
def bulk_upload_students(request):
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = form.cleaned_data['csv_file']
            try:
                decoded_file = csv_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                count = 0
                with transaction.atomic():
                    for row_idx, row in enumerate(reader, start=2):
                        try:
                            branch_id = row['branch_id']
                            try:
                                branch = Branch.objects.get(branch_id=branch_id)
                            except Branch.DoesNotExist:
                                raise ValueError(f"Branch '{branch_id}' does not exist. Please create it first.")
                            
                            drive_id = row['drive_id']
                            try:
                                drive = Drive.objects.get(drive_id=drive_id)
                            except Drive.DoesNotExist:
                                raise ValueError(f"Drive '{drive_id}' does not exist. Please create it first.")

                            placement_status = row.get('placement_status', 'Unplaced')
                            if drive_id.startswith('PD') and placement_status not in ['Unplaced', 'Placed', 'PPO']:
                                raise ValueError(f"Invalid placement_status '{placement_status}' for PD drive.")
                            if drive_id.startswith('SI') and placement_status not in ['Unplaced', 'Summer Internship']:
                                raise ValueError(f"Invalid placement_status '{placement_status}' for SI drive.")

                            student = Student(
                                student_id=row['student_id'],
                                std_name=row['std_name'],
                                branch=branch,
                                drive=drive,
                                cpi=row['cpi'],
                                placement_status=placement_status,
                                offer_letter=row.get('offer_letter', '')
                            )
                            student.full_clean()
                            student.save()
                            count += 1
                        except Exception as e:
                            raise Exception(f"Row {row_idx}: {str(e)}")
                messages.success(request, f'Successfully uploaded {count} students.')
            except Exception as e:
                messages.error(request, f'Error processing CSV - {str(e)}')
    return redirect('students')


@login_required
def download_csv_template(request):
    """Download a sample CSV template for bulk student upload."""
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="student_upload_template.csv"'

    writer = csv.writer(response)
    writer.writerow(['student_id', 'std_name', 'branch_id', 'drive_id', 'cpi', 'placement_status', 'offer_letter'])
    writer.writerow(['STU001', 'John Doe', 'BTECH_CSE', 'SI2025', '8.5', 'Unplaced', ''])
    writer.writerow(['STU002', 'Jane Smith', 'BTECH_EE', 'PD2025', '9.1', 'Placed', 'https://drive.google.com/file/d/example'])

    return response


@login_required
def edit_student(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student updated successfully.')
        else:
            for error in form.errors.values():
                messages.error(request, str(error))
    return redirect('students')


@login_required
def delete_student(request, student_id):
    if request.method == 'POST':
        student = get_object_or_404(Student, pk=student_id)
        name = student.std_name
        student.delete()
        messages.success(request, f'Student "{name}" deleted successfully.')
    return redirect('students')


@login_required
def export_students_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name', 'Branch', 'Drive', 'CPI', 'Placement Status', 'Company', 'Profile Name', 'CTC (LPA)', 'Stipend', 'Offer Letter'])
    
    for std in Student.objects.all():
        writer.writerow([
            std.student_id, std.std_name, std.branch.branch_name, 
            std.drive.drive_name, std.cpi, std.placement_status, 
            std.company.cmp_name if std.company else '', 
            std.profile.profile_name if std.profile else '',
            std.profile.ctc if std.profile else '',
            std.profile.stipend if std.profile else '',
            std.offer_letter or ''
        ])
    return response


@login_required
def companies_list(request):
    companies = Company.objects.all()
    context = {
        'companies': companies,
        'drives': Drive.objects.filter(status='active'),
        'branches': Branch.objects.all(),
        'company_form': CompanyForm(),
        'profile_form': ProfileForm()
    }
    return render(request, 'portal/companies.html', context)


@login_required
def add_company(request):
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company added successfully.')
    return redirect('companies')


@login_required
def edit_company(request, cmp_id):
    cmp = get_object_or_404(Company, pk=cmp_id)
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=cmp)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company updated successfully.')
    return redirect('companies')


@login_required
def delete_company(request, cmp_id):
    if request.method == 'POST':
        cmp = get_object_or_404(Company, pk=cmp_id)
        cmp.delete()
        messages.success(request, 'Company deleted.')
    return redirect('companies')


@login_required
def add_profile(request):
    if request.method == 'POST':
        form = ProfileForm(request.POST)
        cmp_id = request.POST.get('cmp_id')
        cmp = get_object_or_404(Company, pk=cmp_id)
        if form.is_valid():
            profile = form.save(commit=False)
            profile.cmp = cmp
            profile.save()
            form.save_m2m()
            messages.success(request, 'Profile added successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    return redirect('companies')


@login_required
def edit_profile(request, profile_id):
    profile = get_object_or_404(Profile, pk=profile_id)
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    return redirect('companies')


@login_required
def delete_profile(request, profile_id):
    if request.method == 'POST':
        profile = get_object_or_404(Profile, pk=profile_id)
        profile.delete()
        messages.success(request, 'Profile deleted.')
    return redirect('companies')


@login_required
def analytics(request):
    drives = Drive.objects.filter(status='active').order_by('-drive_year')
    drive_stats = []
    for drive in drives:
        is_si = drive.drive_id.startswith('SI')
        students = Student.all_objects.filter(drive=drive)
        total    = students.count()
        placed   = students.filter(placement_status='Placed').count()
        ppo      = students.filter(placement_status='PPO').count()
        si       = students.filter(placement_status='Summer Internship').count()
        unplaced = students.filter(placement_status='Unplaced').count()

        # For CTC stats: SI drives use Summer Internship students; PD drives use Placed+PPO
        if is_si:
            ctc_qs = students.filter(placement_status='Summer Internship', profile__isnull=False).select_related('profile')
            primary_count = si
        else:
            ctc_qs = students.filter(placement_status__in=['Placed', 'PPO'], profile__isnull=False).select_related('profile')
            primary_count = placed + ppo

        if is_si:
            ctcs = [float(s.profile.stipend) / 1000.0 for s in ctc_qs if s.profile and s.profile.stipend]
        else:
            ctcs = [float(s.profile.ctc) for s in ctc_qs if s.profile and s.profile.ctc]
        avg_ctc = round(sum(ctcs) / len(ctcs), 2) if ctcs else 0
        sorted_ctcs = sorted(ctcs)
        n = len(sorted_ctcs)
        median_ctc = round((sorted_ctcs[n//2-1]+sorted_ctcs[n//2])/2 if n%2==0 else sorted_ctcs[n//2], 2) if sorted_ctcs else 0
        highest_ctc = round(max(ctcs), 2) if ctcs else 0
        overall_pct = round(primary_count / total * 100, 1) if total else 0

        branch_data = []
        for branch in drive.branches.all().order_by('branch_name'):
            b_stu = students.filter(branch=branch)
            b_total = b_stu.count()
            b_placed = b_stu.filter(placement_status='Placed').count()
            b_ppo = b_stu.filter(placement_status='PPO').count()
            b_si = b_stu.filter(placement_status='Summer Internship').count()
            b_unplaced = b_stu.filter(placement_status='Unplaced').count()

            if is_si:
                b_ctc_qs = b_stu.filter(placement_status='Summer Internship', profile__isnull=False).select_related('profile')
                b_primary = b_si
            else:
                b_ctc_qs = b_stu.filter(placement_status__in=['Placed','PPO'], profile__isnull=False).select_related('profile')
                b_primary = b_placed + b_ppo

            if is_si:
                b_ctcs = [float(s.profile.stipend) / 1000.0 for s in b_ctc_qs if s.profile and s.profile.stipend]
            else:
                b_ctcs = [float(s.profile.ctc) for s in b_ctc_qs if s.profile and s.profile.ctc]
            b_avg = round(sum(b_ctcs)/len(b_ctcs),2) if b_ctcs else 0
            b_sorted = sorted(b_ctcs)
            bn = len(b_sorted)
            b_median = round((b_sorted[bn//2-1]+b_sorted[bn//2])/2 if bn%2==0 else b_sorted[bn//2],2) if b_sorted else 0
            b_highest = round(max(b_ctcs),2) if b_ctcs else 0
            b_pct = round(b_primary / b_total * 100, 1) if b_total else 0

            branch_data.append({
                'branch_name': branch.branch_name, 'branch_id': branch.branch_id,
                'total': b_total, 'placed': b_placed, 'ppo': b_ppo,
                'si': b_si, 'unplaced': b_unplaced,
                'placed_total': b_placed + b_ppo,
                'primary_count': b_primary,
                'avg_ctc': b_avg, 'median_ctc': b_median, 'highest_ctc': b_highest,
                'placement_pct': b_pct,
                'opened_profiles': Profile.objects.filter(cmp__drive=drive, branches=branch).count(),
                'opened_profiles_list': Profile.objects.filter(cmp__drive=drive, branches=branch).select_related('cmp'),
            })

        drive_stats.append({
            'drive': drive, 'is_si': is_si,
            'total': total, 'placed': placed, 'ppo': ppo,
            'si': si, 'unplaced': unplaced,
            'placed_total': placed + ppo, 'primary_count': primary_count,
            'avg_ctc': avg_ctc, 'median_ctc': median_ctc, 'highest_ctc': highest_ctc,
            'overall_pct': overall_pct, 'branches': branch_data,
        })
    return render(request, 'portal/analytics.html', {'drive_stats': drive_stats})


@login_required
def archive(request):
    # Get all drives that are inactive
    archived_drives = Drive.objects.filter(status='inactive')
    context = {'archived_drives': archived_drives}
    return render(request, 'portal/archive.html', context)


@login_required
def restore_drive(request, drive_id):
    if request.method == 'POST':
        drive = get_object_or_404(Drive, pk=drive_id)
        try:
            drive.status = 'active'
            drive.clean()
            drive.save()
            messages.success(request, f'Drive {drive.drive_name} restored successfully.')
        except ValidationError as e:
            messages.error(request, e.message)
    return redirect('archive')


@login_required
def export_archive_csv(request, drive_id):
    drive = get_object_or_404(Drive, pk=drive_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="archive_{drive.drive_name}.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Name', 'Branch', 'CPI', 'Placement Status', 'Company', 'Profile Name', 'CTC (LPA)', 'Stipend', 'Offer Letter'])
    
    # Bypass ActiveDriveManager for archive export
    students = Student.all_objects.filter(drive=drive)
    for std in students:
        writer.writerow([
            std.student_id, std.std_name, std.branch.branch_name, 
            std.cpi, std.placement_status, 
            std.company.cmp_name if std.company else '', 
            std.profile.profile_name if std.profile else '',
            std.profile.ctc if std.profile else '',
            std.profile.stipend if std.profile else '',
            std.offer_letter or ''
        ])
    return response


@csrf_exempt
def rag_query(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            query = data.get('query')
            
            # Aggregate stats
            total_students = Student.all_objects.count()
            placed_students = Student.all_objects.exclude(placement_status='Unplaced').count()
            
            context_str = f"Total Students: {total_students}. Placed: {placed_students}."
            
            # Simple RAG integration using anthropic if available
            api_key = os.environ.get('ANTHROPIC_API_KEY')
            if Anthropic and api_key:
                client = Anthropic(api_key=api_key)
                response = client.messages.create(
                    model="claude-3-5-sonnet-20241022",
                    max_tokens=500,
                    system="You are a helpful placement analytics assistant.",
                    messages=[
                        {"role": "user", "content": f"Context: {context_str}\n\nQuestion: {query}"}
                    ]
                )
                answer = response.content[0].text
            else:
                # Mock response if no key
                answer = f"I see your question: '{query}'. Based on my context, we have {total_students} total students and {placed_students} are placed. (Please configure ANTHROPIC_API_KEY for full AI responses)."
                
            return JsonResponse({'response': answer})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid method'}, status=400)


@login_required
def analytics_data(request):
    data = {}
    for d in Drive.objects.filter(status='active'):
        placed = d.students.filter(placement_status='Placed').count()
        ppo = d.students.filter(placement_status='PPO').count()
        si = d.students.filter(placement_status='Summer Internship').count()
        unplaced = d.students.filter(placement_status='Unplaced').count()
        
        labels = []
        counts = []
        colors = []
        
        if placed > 0:
            labels.append('Placed')
            counts.append(placed)
            colors.append('#28a745')
        if ppo > 0:
            labels.append('PPO')
            counts.append(ppo)
            colors.append('#17a2b8')
        if si > 0:
            labels.append('Summer Internship')
            counts.append(si)
            colors.append('#ffc107')
        if unplaced > 0 or len(labels) == 0:
            labels.append('Unplaced')
            counts.append(unplaced)
            colors.append('#dc3545')
        
        data[d.drive_id] = {
            'labels': labels,
            'datasets': [{
                'data': counts,
                'backgroundColor': colors
            }]
        }
        
    return JsonResponse({'overview': data})


@login_required
def get_profiles_for_drive(request, drive_id):
    profiles = Profile.objects.filter(cmp__drive_id=drive_id)
    return JsonResponse({'profiles': list(profiles.values('profile_id', 'profile_name', 'cmp__cmp_name'))})


@login_required
def get_profiles_for_company(request, cmp_id):
    profiles = Profile.objects.filter(cmp_id=cmp_id)
    return JsonResponse({'profiles': list(profiles.values('profile_id', 'profile_name', 'ctc', 'stipend'))})


@login_required
@csrf_exempt
def text_to_sql_report(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_prompt = data.get('prompt')
            
            if not user_prompt:
                return JsonResponse({"error": "No prompt provided"}, status=400)

            # Local import to prevent circular dependencies
            from .services import generate_excel_from_prompt
            excel_file = generate_excel_from_prompt(user_prompt)
            
            response = HttpResponse(
                excel_file.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="Placement_Report.xlsx"'
            return response

        except ValueError as ve:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": str(ve)}, status=400)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": f"An error occurred: {str(e)}"}, status=500)
            
    return JsonResponse({"error": "Only POST requests are allowed"}, status=405)


# ── Profile Rounds ──
@login_required
def profile_rounds(request, profile_id):
    profile = get_object_or_404(Profile, pk=profile_id)
    rounds = profile.rounds.all().order_by('round_number')
    
    # Pre-populate round student choices for manual adding
    drive = profile.cmp.drive
    drive_students = Student.objects.filter(drive=drive)
    if profile.branches.exists():
        drive_students = drive_students.filter(branch__in=profile.branches.all())
    
    for r in rounds:
        r.existing_student_ids = set(r.round_students.values_list('student_id', flat=True))
        if r.round_number == 1:
            r.eligible_students = drive_students.exclude(student_id__in=r.existing_student_ids).order_by('std_name')
        else:
            prev_round = rounds.filter(round_number=r.round_number - 1).first()
            if prev_round:
                prev_student_ids = prev_round.round_students.values_list('student_id', flat=True)
                r.eligible_students = Student.objects.filter(student_id__in=prev_student_ids).exclude(student_id__in=r.existing_student_ids).order_by('std_name')
            else:
                r.eligible_students = Student.objects.none()

    context = {
        'profile': profile,
        'rounds': rounds,
        'round_form': InterviewRoundForm(),
        'csv_form': CSVUploadForm(),
    }
    return render(request, 'portal/profile_rounds.html', context)


@login_required
def add_round(request, profile_id):
    profile = get_object_or_404(Profile, pk=profile_id)
    if request.method == 'POST':
        form = InterviewRoundForm(request.POST)
        if form.is_valid():
            round_obj = form.save(commit=False)
            round_obj.profile = profile
            max_num = profile.rounds.aggregate(Max('round_number'))['round_number__max'] or 0
            round_obj.round_number = max_num + 1
            round_obj.save()
            messages.success(request, f"Round '{round_obj.round_name}' added successfully.")
        else:
            for error in form.errors.values():
                messages.error(request, str(error))
    return redirect('profile_rounds', profile_id=profile_id)


@login_required
def edit_round(request, round_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    if request.method == 'POST':
        form = InterviewRoundForm(request.POST, instance=round_obj)
        if form.is_valid():
            form.save()
            messages.success(request, f"Round updated successfully.")
        else:
            for error in form.errors.values():
                messages.error(request, str(error))
    return redirect('profile_rounds', profile_id=round_obj.profile.profile_id)


@login_required
def delete_round(request, round_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    profile = round_obj.profile
    if round_obj.round_type == 'applications':
        messages.error(request, "The 'Applications' round cannot be deleted.")
        return redirect('profile_rounds', profile_id=profile.profile_id)
    
    if request.method == 'POST':
        round_number = round_obj.round_number
        round_obj.delete()
        
        # Re-number subsequent rounds
        subsequent_rounds = profile.rounds.filter(round_number__gt=round_number)
        for r in subsequent_rounds:
            r.round_number -= 1
            r.save()
            
        messages.success(request, "Round deleted successfully.")
    return redirect('profile_rounds', profile_id=profile.profile_id)


@csrf_exempt
@login_required
def reorder_rounds(request, profile_id):
    profile = get_object_or_404(Profile, pk=profile_id)
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            round_ids = data.get('round_ids', [])
            
            with transaction.atomic():
                for idx, r_id in enumerate(round_ids, start=1):
                    InterviewRound.objects.filter(pk=r_id, profile=profile).update(round_number=idx)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)


@login_required
def upload_round_students_csv(request, round_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = form.cleaned_data['csv_file']
            try:
                decoded_file = csv_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                
                count = 0
                errors = []
                
                profile = round_obj.profile
                drive = profile.cmp.drive
                
                if round_obj.round_number == 1:
                    eligible_qs = Student.objects.filter(drive=drive)
                    if profile.branches.exists():
                        eligible_qs = eligible_qs.filter(branch__in=profile.branches.all())
                    eligible_ids = set(eligible_qs.values_list('student_id', flat=True))
                else:
                    prev_round = profile.rounds.filter(round_number=round_obj.round_number - 1).first()
                    if prev_round:
                        eligible_ids = set(prev_round.round_students.values_list('student_id', flat=True))
                    else:
                        eligible_ids = set()

                existing_ids = set(round_obj.round_students.values_list('student_id', flat=True))

                with transaction.atomic():
                    for row_idx, row in enumerate(reader, start=2):
                        student_id = row.get('student_id')
                        if not student_id:
                            errors.append(f"Row {row_idx}: Missing student_id.")
                            continue
                        
                        student_id = student_id.strip()
                        if student_id in existing_ids:
                            continue
                        
                        if student_id not in eligible_ids:
                            if round_obj.round_number == 1:
                                errors.append(f"Row {row_idx}: Student '{student_id}' is not eligible/in drive for this profile.")
                            else:
                                errors.append(f"Row {row_idx}: Student '{student_id}' was not in previous round.")
                            continue
                        
                        student = Student.objects.get(pk=student_id)
                        RoundStudent.objects.create(round=round_obj, student=student)
                        count += 1
                
                if errors:
                    messages.warning(request, f"Uploaded {count} students. Issues: " + "; ".join(errors[:5]))
                else:
                    messages.success(request, f"Successfully uploaded {count} students.")
            except Exception as e:
                messages.error(request, f"Error processing CSV: {str(e)}")
    return redirect('profile_rounds', profile_id=round_obj.profile.profile_id)


@login_required
def add_round_student(request, round_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        single_student_id = request.POST.get('student_id')
        if single_student_id and single_student_id not in student_ids:
            student_ids.append(single_student_id)
            
        student_ids = [s_id.strip() for s_id in student_ids if s_id.strip()]
        if not student_ids:
            messages.error(request, "Please select at least one student.")
            return redirect('profile_rounds', profile_id=round_obj.profile.profile_id)
            
        profile = round_obj.profile
        drive = profile.cmp.drive
        
        # Precautionary check of eligible students
        if round_obj.round_number == 1:
            eligible_qs = Student.objects.filter(drive=drive)
            if profile.branches.exists():
                eligible_qs = eligible_qs.filter(branch__in=profile.branches.all())
            eligible_ids = set(eligible_qs.values_list('student_id', flat=True))
        else:
            prev_round = profile.rounds.filter(round_number=round_obj.round_number - 1).first()
            if prev_round:
                eligible_ids = set(prev_round.round_students.values_list('student_id', flat=True))
            else:
                eligible_ids = set()

        existing_ids = set(round_obj.round_students.values_list('student_id', flat=True))
        
        added_count = 0
        skipped_count = 0
        error_messages = []
        
        with transaction.atomic():
            for student_id in student_ids:
                if student_id in existing_ids:
                    skipped_count += 1
                    continue
                if student_id not in eligible_ids:
                    if round_obj.round_number == 1:
                        error_messages.append(f"Student {student_id} is not in drive or branch is not eligible.")
                    else:
                        error_messages.append(f"Student {student_id} was not in previous round.")
                    continue
                
                try:
                    student = Student.objects.get(pk=student_id)
                    RoundStudent.objects.create(round=round_obj, student=student)
                    added_count += 1
                except Student.DoesNotExist:
                    error_messages.append(f"Student {student_id} does not exist.")
                    
        if added_count > 0:
            messages.success(request, f"Successfully added {added_count} student(s) to this round.")
        if skipped_count > 0:
            messages.info(request, f"{skipped_count} student(s) were already in this round.")
        if error_messages:
            messages.warning(request, "Some students could not be added: " + "; ".join(error_messages[:5]))
            
    return redirect('profile_rounds', profile_id=round_obj.profile.profile_id)


@login_required
def remove_round_student(request, round_id, student_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    student = get_object_or_404(Student, pk=student_id)
    if request.method == 'POST':
        RoundStudent.objects.filter(round=round_obj, student=student).delete()
        subsequent_rounds = InterviewRound.objects.filter(
            profile=round_obj.profile,
            round_number__gt=round_obj.round_number
        )
        RoundStudent.objects.filter(round__in=subsequent_rounds, student=student).delete()
        messages.success(request, f"Removed student {student.std_name} from this round and subsequent rounds.")
    return redirect('profile_rounds', profile_id=round_obj.profile.profile_id)


@login_required
def toggle_round_final(request, round_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    profile = round_obj.profile
    if request.method == 'POST':
        if not round_obj.is_final:
            with transaction.atomic():
                profile.rounds.update(is_final=False)
                round_obj.is_final = True
                round_obj.save()
                
                # Reset previous selections for this profile
                Student.all_objects.filter(profile=profile).update(
                    placement_status='Unplaced',
                    company=None,
                    profile=None
                )
                
                count = 0
                for rs in round_obj.round_students.select_related('student'):
                    student = rs.student
                    if student.drive.is_si:
                        student.placement_status = 'Summer Internship'
                    else:
                        student.placement_status = 'Placed'
                    student.company = profile.cmp
                    student.profile = profile
                    student.save()
                    count += 1
            messages.success(request, f"Round marked as Final Results. Synced {count} students to placement status.")
        else:
            with transaction.atomic():
                round_obj.is_final = False
                round_obj.save()
                Student.all_objects.filter(profile=profile).update(
                    placement_status='Unplaced',
                    company=None,
                    profile=None
                )
            messages.success(request, f"Round unmarked as Final Results. Placement status reset for these students.")
            
    return redirect('profile_rounds', profile_id=profile.profile_id)


@login_required
def export_round_students(request, round_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    students = Student.objects.filter(interview_rounds__round=round_obj).select_related('branch')
    
    data = []
    for s in students:
        data.append({
            'Student ID': s.student_id,
            'Name': s.std_name,
            'Branch': s.branch.branch_name,
            'CPI': s.cpi
        })
        
    df = pd.DataFrame(data)
    
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Shortlist')
        
    excel_buffer.seek(0)
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = round_obj.round_name.replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="{safe_name}_Shortlist.xlsx"'
    return response


@login_required
def export_profile_journey_excel(request, profile_id):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    profile = get_object_or_404(Profile, pk=profile_id)
    rounds = list(profile.rounds.all().order_by('round_number'))
    
    if not rounds:
        messages.error(request, "This profile has no rounds.")
        return redirect('companies')
        
    first_round = rounds[0]
    # Get all students who applied (i.e. participated in the first round)
    applied_students = Student.objects.filter(interview_rounds__round=first_round).select_related('branch').order_by('student_id')
    
    # Other rounds (excluding the first one, e.g. "Applications")
    other_rounds = rounds[1:]
    
    # Get all RoundStudent mapping for database query optimization
    round_student_pairs = set(
        RoundStudent.objects.filter(round__profile=profile).values_list('round_id', 'student_id')
    )
    
    # Determine which round is the final round
    final_round = next((r for r in reversed(rounds) if r.is_final), rounds[-1])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Selection Status"
    
    # Build headers
    headers = ['Sr No.', 'Student ID', 'Student Name', 'Branch', 'Student CPI']
    for r in other_rounds:
        headers.append(r.round_name)
    
    ws.append(headers)
    
    # Define styles - Yellow for header, Green/Red for rounds
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    header_font = Font(name='Calibri', size=11, bold=True)
    data_font = Font(name='Calibri', size=11, bold=False)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Apply header formatting
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = yellow_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx != 3 and col_idx != 4 else 'left', vertical='center')
        
    # Add data rows
    for idx, s in enumerate(applied_students, 1):
        row_data = [
            idx,
            s.student_id,
            s.std_name,
            s.branch.branch_name,
            float(s.cpi) if s.cpi else 0.0
        ]
        # Append empty cells for rounds
        for _ in other_rounds:
            row_data.append(None)
        
        ws.append(row_data)
        row_idx = idx + 1
        
        # Apply standard styles for first 5 cells
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [1, 2, 5]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
        # Apply round status formatting (Green if in round, Red if not)
        col_offset = 6
        for r_idx, r in enumerate(other_rounds):
            cell = ws.cell(row=row_idx, column=col_offset + r_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if (r.round_id, s.student_id) in round_student_pairs:
                cell.fill = green_fill
            else:
                cell.fill = red_fill
            
    # Set explicit column widths to match user's template
    col_widths = {
        'A': 12,
        'B': 14,
        'C': 16,
        'D': 14,
        'E': 15
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    for r_idx in range(len(other_rounds)):
        col_letter = get_column_letter(6 + r_idx)
        ws.column_dimensions[col_letter].width = 16
    
    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Send file
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = profile.profile_name.replace(" ", "_")
    company_name = profile.cmp.cmp_name.replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="{company_name}_{safe_name}_status.xlsx"'
    return response


@login_required
def toggle_interview_shortlist(request, round_id):
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    if request.method == 'POST':
        round_obj.is_interview_shortlist = not round_obj.is_interview_shortlist
        round_obj.save()
        if round_obj.is_interview_shortlist:
            messages.success(request, f"'{round_obj.round_name}' marked as Interview Shortlist round.")
        else:
            messages.success(request, f"'{round_obj.round_name}' unmarked as Interview Shortlist round.")
    return redirect('profile_rounds', profile_id=round_obj.profile.profile_id)


@login_required
def download_round_csv_template(request, round_id):
    """Download a sample CSV template showing the expected format for round student uploads."""
    round_obj = get_object_or_404(InterviewRound, pk=round_id)
    response = HttpResponse(content_type='text/csv')
    safe_name = round_obj.round_name.replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="{safe_name}_upload_template.csv"'
    writer = csv.writer(response)
    writer.writerow(['student_id'])
    writer.writerow(['EXAMPLE_STU001'])
    writer.writerow(['EXAMPLE_STU002'])
    return response


@login_required
def student_detail(request, student_id):
    student = get_object_or_404(Student.all_objects, pk=student_id)
    
    round_students = RoundStudent.objects.filter(student=student).select_related(
        'round__profile__cmp', 'round__profile'
    )
    
    profile_rounds_map = {}
    for rs in round_students:
        prof = rs.round.profile
        if prof.profile_id not in profile_rounds_map:
            profile_rounds_map[prof.profile_id] = {
                'profile': prof,
                'rounds_present': [],
            }
        profile_rounds_map[prof.profile_id]['rounds_present'].append(rs.round)
        
    journey = []
    total_applications = 0
    total_shortlists = 0
    total_selections = 0
    
    for prof_id, info in profile_rounds_map.items():
        prof = info['profile']
        rounds_present = info['rounds_present']
        
        all_rounds = list(prof.rounds.all().order_by('round_number'))
        rounds_present_ids = {r.round_id for r in rounds_present}
        max_round_num = max([r.round_number for r in rounds_present], default=0)
        
        is_applied = any(r.round_type == 'applications' for r in rounds_present)
        if is_applied:
            total_applications += 1
            
        is_shortlisted = any(r.is_interview_shortlist and r.round_id in rounds_present_ids for r in all_rounds)
        if is_shortlisted:
            total_shortlists += 1
            
        is_selected = any(r.is_final and r.round_id in rounds_present_ids for r in all_rounds)
        if is_selected:
            total_selections += 1
            
        round_statuses = []
        is_ended = any(r.is_final for r in all_rounds)
        for r in all_rounds:
            cleared = r.round_id in rounds_present_ids
            if cleared:
                r_status = 'Cleared'
            else:
                if is_ended:
                    r_status = 'Eliminated'
                else:
                    if r.round_number > max_round_num:
                        has_subsequent_students = RoundStudent.objects.filter(
                            round__profile=prof,
                            round__round_number__gte=r.round_number
                        ).exists()
                        if has_subsequent_students:
                            r_status = 'Eliminated'
                        else:
                            r_status = 'Pending'
                    else:
                        r_status = 'Eliminated'
            
            round_statuses.append({
                'round': r,
                'cleared': cleared,
                'status': r_status
            })
            
        journey.append({
            'profile': prof,
            'round_statuses': round_statuses,
            'is_selected': is_selected,
            'is_ended': is_ended,
            'max_round_num': max_round_num,
            'cleared_count': len(rounds_present)
        })
        
    context = {
        'student': student,
        'journey': journey,
        'total_applications': total_applications,
        'total_shortlists': total_shortlists,
        'total_selections': total_selections,
    }
    return render(request, 'portal/student_detail.html', context)


@login_required
def export_student_detail(request, student_id):
    student = get_object_or_404(Student.all_objects, pk=student_id)
    round_students = RoundStudent.objects.filter(student=student).select_related(
        'round__profile__cmp', 'round__profile'
    )
    
    profile_rounds_map = {}
    for rs in round_students:
        prof = rs.round.profile
        if prof.profile_id not in profile_rounds_map:
            profile_rounds_map[prof.profile_id] = []
        profile_rounds_map[prof.profile_id].append(rs.round)
        
    total_applications = 0
    total_shortlists = 0
    total_selections = 0
    
    journey_rows = []
    for prof_id, rounds in profile_rounds_map.items():
        prof = rounds[0].profile
        all_rounds = list(prof.rounds.all().order_by('round_number'))
        rounds_present_ids = {r.round_id for r in rounds}
        
        is_applied = any(r.round_type == 'applications' for r in rounds)
        if is_applied:
            total_applications += 1
            
        is_shortlisted = any(r.is_interview_shortlist and r.round_id in rounds_present_ids for r in all_rounds)
        if is_shortlisted:
            total_shortlists += 1
            
        is_selected = any(r.is_final and r.round_id in rounds_present_ids for r in all_rounds)
        if is_selected:
            total_selections += 1
            
        is_ended = any(r.is_final for r in all_rounds)
        max_round = max([r for r in rounds], key=lambda x: x.round_number, default=None)
        highest_round_cleared = max_round.round_name if max_round else '-'
        
        if is_selected:
            status_str = "Selected"
        else:
            if is_ended:
                status_str = "Ended"
            else:
                status_str = "In Progress"
                
        journey_rows.append({
            'Company': prof.cmp.cmp_name,
            'Profile': prof.profile_name,
            'Offer Type': prof.get_offer_type_display() if prof.offer_type else '-',
            'CTC (LPA)': prof.ctc,
            'Stipend': prof.stipend or '-',
            'Rounds Cleared': len(rounds),
            'Highest Round Cleared': highest_round_cleared,
            'Status': status_str
        })
        
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        summary_df = pd.DataFrame([{
            'Student ID': student.student_id,
            'Name': student.std_name,
            'Branch': student.branch.branch_name,
            'CPI': student.cpi,
            'Placement Status': student.placement_status,
            'Company': student.company.cmp_name if student.company else '-',
            'Profile': student.profile.profile_name if student.profile else '-',
            'Total Applications': total_applications,
            'Shortlisted for Interviews': total_shortlists,
            'Total Selections': total_selections
        }])
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        
        journey_df = pd.DataFrame(journey_rows)
        journey_df.to_excel(writer, index=False, sheet_name='Interview Journey')
        
    excel_buffer.seek(0)
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = student.std_name.replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="Journey_{safe_name}.xlsx"'
    return response

            
    return JsonResponse({"error": "Only POST requests are allowed"}, status=405)