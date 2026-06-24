from django.test import TestCase
from django.core.exceptions import ValidationError
from django.contrib.auth.models import User
from django.urls import reverse
from portal.models import Branch, Drive, Student, Company, Profile, InterviewRound, RoundStudent
from portal.forms import StudentForm


class StudentOfferLetterTests(TestCase):
    def setUp(self):
        # Create Branch
        self.branch = Branch.objects.create(
            branch_id="BTECH_CSE",
            branch_name="Computer Science"
        )
        # Create Drives
        self.active_pd_drive = Drive.objects.create(
            drive_id="PD2025",
            drive_name="Placement Drive 2025",
            drive_year=2025,
            status="active"
        )
        self.active_pd_drive.branches.add(self.branch)
        
        self.active_si_drive = Drive.objects.create(
            drive_id="SI2025",
            drive_name="Summer Internship 2025",
            drive_year=2025,
            status="active"
        )
        self.active_si_drive.branches.add(self.branch)

        # Create Company
        self.company = Company.objects.create(
            cmp_name="Google",
            drive=self.active_pd_drive
        )

        # Create Profile
        self.profile = Profile.objects.create(
            cmp=self.company,
            profile_name="Software Engineer",
            ctc=15.0,
            stipend=50000.0,
            offer_type="J"
        )

    def test_unplaced_student_cannot_have_offer_letter(self):
        # Model level validation
        student = Student(
            student_id="STU001",
            std_name="John Doe",
            branch=self.branch,
            drive=self.active_pd_drive,
            cpi=8.5,
            placement_status="Unplaced",
            offer_letter="https://drive.google.com/file/d/12345"
        )
        with self.assertRaises(ValidationError):
            student.full_clean()

    def test_placed_student_can_have_offer_letter(self):
        student = Student(
            student_id="STU001",
            std_name="John Doe",
            branch=self.branch,
            drive=self.active_pd_drive,
            cpi=8.5,
            placement_status="Placed",
            company=self.company,
            profile=self.profile,
            offer_letter="https://drive.google.com/file/d/12345"
        )
        # Should clean successfully
        student.full_clean()
        student.save()
        
        saved = Student.all_objects.get(pk="STU001")
        self.assertEqual(saved.offer_letter, "https://drive.google.com/file/d/12345")

    def test_form_clears_offer_letter_if_unplaced(self):
        # When sending Unplaced status, form should clear offer letter
        form_data = {
            'student_id': "STU002",
            'std_name': "Jane Smith",
            'branch': self.branch.branch_id,
            'drive': self.active_pd_drive.drive_id,
            'cpi': 9.0,
            'placement_status': "Unplaced",
            'company': "",
            'profile': "",
            'offer_letter': "https://drive.google.com/file/d/12345"
        }
        form = StudentForm(data=form_data)
        self.assertTrue(form.is_valid())
        self.assertEqual(form.cleaned_data['offer_letter'], "")

    def test_form_retains_offer_letter_if_placed(self):
        form_data = {
            'student_id': "STU003",
            'std_name': "Bob Johnson",
            'branch': self.branch.branch_id,
            'drive': self.active_pd_drive.drive_id,
            'cpi': 7.5,
            'placement_status': "Placed",
            'company': self.company.cmp_id,
            'profile': self.profile.profile_id,
            'offer_letter': "https://drive.google.com/file/d/99999"
        }
        form = StudentForm(data=form_data)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['offer_letter'], "https://drive.google.com/file/d/99999")


class ProfileBranchTests(TestCase):
    def setUp(self):
        self.branch_cse = Branch.objects.create(
            branch_id="BTECH_CSE",
            branch_name="Computer Science"
        )
        self.branch_ece = Branch.objects.create(
            branch_id="BTECH_ECE",
            branch_name="Electronics"
        )
        self.active_drive = Drive.objects.create(
            drive_id="PD2025",
            drive_name="Placement Drive 2025",
            drive_year=2025,
            status="active"
        )
        self.active_drive.branches.add(self.branch_cse, self.branch_ece)

        self.company = Company.objects.create(
            cmp_name="Meta",
            drive=self.active_drive
        )

    def test_profile_can_have_multiple_branches(self):
        profile = Profile.objects.create(
            cmp=self.company,
            profile_name="Product Manager",
            ctc=20.0
        )
        profile.branches.add(self.branch_cse, self.branch_ece)
        
        self.assertEqual(profile.branches.count(), 2)
        self.assertIn(self.branch_cse, profile.branches.all())
        self.assertIn(self.branch_ece, profile.branches.all())


class ProfileOfferTypeValidationTests(TestCase):
    def setUp(self):
        self.branch = Branch.objects.create(
            branch_id="BTECH_CSE",
            branch_name="Computer Science"
        )
        self.si_drive = Drive.objects.create(
            drive_id="SI2025",
            drive_name="Summer Internship 2025",
            drive_year=2025,
            status="active"
        )
        self.si_drive.branches.add(self.branch)
        
        self.pd_drive = Drive.objects.create(
            drive_id="PD2025",
            drive_name="Placement Drive 2025",
            drive_year=2025,
            status="active"
        )
        self.pd_drive.branches.add(self.branch)

        self.si_company = Company.objects.create(
            cmp_name="Google SI",
            drive=self.si_drive
        )
        self.pd_company = Company.objects.create(
            cmp_name="Google PD",
            drive=self.pd_drive
        )

    def test_si_drive_only_allows_si_offer_type(self):
        # Invalid offer type J in SI drive
        profile_invalid = Profile(
            cmp=self.si_company,
            profile_name="SWE Intern",
            ctc=12.0,
            offer_type="J"
        )
        with self.assertRaises(ValidationError):
            profile_invalid.full_clean()

        # Valid offer type SI in SI drive
        profile_valid = Profile(
            cmp=self.si_company,
            profile_name="SWE Intern",
            ctc=12.0,
            offer_type="SI"
        )
        profile_valid.full_clean()  # Should not raise any error

    def test_pd_drive_does_not_allow_si_offer_type(self):
        # Invalid offer type SI in PD drive
        profile_invalid = Profile(
            cmp=self.pd_company,
            profile_name="Full Time SWE",
            ctc=25.0,
            offer_type="SI"
        )
        with self.assertRaises(ValidationError):
            profile_invalid.full_clean()

        # Valid offer type J in PD drive
        profile_valid = Profile(
            cmp=self.pd_company,
            profile_name="Full Time SWE",
            ctc=25.0,
            offer_type="J"
        )
        profile_valid.full_clean()  # Should not raise any error


class TextToSQLTests(TestCase):
    def test_safe_query_validation(self):
        from portal.services import is_safe_query
        self.assertTrue(is_safe_query("SELECT * FROM student"))
        self.assertTrue(is_safe_query("SELECT std_name, cpi FROM student WHERE cpi > 8.0"))
        
        self.assertFalse(is_safe_query("INSERT INTO student (student_id) VALUES ('1')"))
        self.assertFalse(is_safe_query("UPDATE student SET cpi = 10.0"))
        self.assertFalse(is_safe_query("DELETE FROM student"))
        self.assertFalse(is_safe_query("DROP TABLE student"))

    def test_endpoint_requires_login(self):
        response = self.client.post('/generate-report/', {'prompt': 'Show students'}, content_type='application/json')
        # Should redirect to login page (302 status code)
        self.assertEqual(response.status_code, 302)


class InterviewRoundTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='testadmin', password='password')
        from portal.models import AdminProfile
        AdminProfile.objects.create(user=self.user, role='super_admin')
        self.client.force_login(self.user)
        
        self.branch = Branch.objects.create(
            branch_id="BTECH_CSE",
            branch_name="Computer Science"
        )
        self.drive = Drive.objects.create(
            drive_id="PD2025",
            drive_name="Placement Drive 2025",
            drive_year=2025,
            status="active"
        )
        self.drive.branches.add(self.branch)
        
        self.company = Company.objects.create(
            cmp_name="Meta",
            drive=self.drive
        )
        # Creating profile should auto-create Applications round
        self.profile = Profile.objects.create(
            cmp=self.company,
            profile_name="Software Engineer",
            ctc=20.0,
            offer_type="J"
        )
        self.profile.branches.add(self.branch)
        
        # Create some students
        self.student1 = Student.objects.create(
            student_id="STU001",
            std_name="Alice",
            branch=self.branch,
            drive=self.drive,
            cpi=9.0
        )
        self.student2 = Student.objects.create(
            student_id="STU002",
            std_name="Bob",
            branch=self.branch,
            drive=self.drive,
            cpi=8.0
        )

    def test_profile_creation_auto_creates_applications_round(self):
        rounds = self.profile.rounds.all()
        self.assertEqual(rounds.count(), 1)
        app_round = rounds.first()
        self.assertEqual(app_round.round_name, "Applications")
        self.assertEqual(app_round.round_type, "applications")
        self.assertEqual(app_round.round_number, 1)

    def test_add_round_view(self):
        url = reverse('add_round', args=[self.profile.profile_id])
        response = self.client.post(url, {
            'round_name': 'OA Shortlist',
            'round_type': 'oa'
        })
        self.assertEqual(response.status_code, 302)
        
        # print messages if failure
        from django.contrib.messages import get_messages
        msgs = [str(m) for m in get_messages(response.wsgi_request)]
        print("\n[test_add_round_view] Messages:", msgs)
        print("[test_add_round_view] Location:", response.get('Location'))
        
        rounds = self.profile.rounds.all().order_by('round_number')
        self.assertEqual(rounds.count(), 2)
        oa_round = rounds[1]
        self.assertEqual(oa_round.round_name, 'OA Shortlist')
        self.assertEqual(oa_round.round_type, 'oa')
        self.assertEqual(oa_round.round_number, 2)

    def test_add_student_manually_to_round_1_and_round_2(self):
        app_round = self.profile.rounds.first()
        
        # Add Alice to Applications
        url = reverse('add_round_student', args=[app_round.round_id])
        response = self.client.post(url, {'student_id': self.student1.student_id})
        self.assertEqual(response.status_code, 302)
        
        from django.contrib.messages import get_messages
        msgs = [str(m) for m in get_messages(response.wsgi_request)]
        print("\n[test_add_student_manually] Add Round 1 Messages:", msgs)
        
        self.assertTrue(RoundStudent.objects.filter(round=app_round, student=self.student1).exists())
        
        # Create Round 2 (OA Shortlist)
        oa_round = InterviewRound.objects.create(
            profile=self.profile,
            round_number=2,
            round_name="OA Shortlist",
            round_type="oa"
        )
        
        # Try to add Bob to Round 2 (Bob is not in Round 1, so should fail)
        url_2 = reverse('add_round_student', args=[oa_round.round_id])
        response = self.client.post(url_2, {'student_id': self.student2.student_id})
        self.assertEqual(response.status_code, 302)
        self.assertFalse(RoundStudent.objects.filter(round=oa_round, student=self.student2).exists())
        
        # Add Alice to Round 2 (Alice is in Round 1, so should succeed)
        response = self.client.post(url_2, {'student_id': self.student1.student_id})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(RoundStudent.objects.filter(round=oa_round, student=self.student1).exists())

    def test_remove_student_cascades_to_subsequent_rounds(self):
        app_round = self.profile.rounds.first()
        oa_round = InterviewRound.objects.create(
            profile=self.profile,
            round_number=2,
            round_name="OA Shortlist",
            round_type="oa"
        )
        
        # Add Alice to both rounds
        RoundStudent.objects.create(round=app_round, student=self.student1)
        RoundStudent.objects.create(round=oa_round, student=self.student1)
        
        # Remove Alice from Applications (should cascade to OA Shortlist)
        url = reverse('remove_round_student', args=[app_round.round_id, self.student1.student_id])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 302)
        
        self.assertFalse(RoundStudent.objects.filter(student=self.student1).exists())

    def test_toggle_round_final_syncs_students_placement_status(self):
        app_round = self.profile.rounds.first()
        RoundStudent.objects.create(round=app_round, student=self.student1)
        
        # Create results round and add Alice
        results_round = InterviewRound.objects.create(
            profile=self.profile,
            round_number=2,
            round_name="Results",
            round_type="results"
        )
        RoundStudent.objects.create(round=results_round, student=self.student1)
        
        # Toggle results round as final
        url = reverse('toggle_round_final', args=[results_round.round_id])
        response = self.client.post(url)
        self.assertEqual(response.status_code, 302)
        
        # Student status should sync
        self.student1.refresh_from_db()
        self.assertEqual(self.student1.placement_status, "Placed")
        self.assertEqual(self.student1.company, self.company)
        self.assertEqual(self.student1.profile, self.profile)
        
        # Toggle it off
        response = self.client.post(url)
        self.assertEqual(response.status_code, 302)
        self.student1.refresh_from_db()
        self.assertEqual(self.student1.placement_status, "Unplaced")
        self.assertIsNone(self.student1.company)
        self.assertIsNone(self.student1.profile)

    def test_export_profile_journey_excel(self):
        # Add student 1 to applications
        app_round = self.profile.rounds.first()
        RoundStudent.objects.create(round=app_round, student=self.student1)

        # Create technical round and add student 1
        tech_round = InterviewRound.objects.create(
            profile=self.profile,
            round_number=2,
            round_name="Technical Round",
            round_type="technical"
        )
        RoundStudent.objects.create(round=tech_round, student=self.student1)

        # Alice is in Applications and Technical Round.
        url = reverse('export_profile_journey_excel', args=[self.profile.profile_id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Load returned sheet with openpyxl to verify
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.title, "Selection Status")
        
        # Headers should be: Sr No., Student ID, Student Name, Branch, Student CPI, Technical Round
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ['Sr No.', 'Student ID', 'Student Name', 'Branch', 'Student CPI', 'Technical Round'])
        
        # Row 2 should be Alice
        row2 = [cell.value for cell in ws[2]]
        self.assertEqual(row2[0], 1)
        self.assertEqual(row2[1], 'STU001')
        self.assertEqual(row2[2], 'Alice')
        self.assertEqual(row2[3], 'Computer Science')
        self.assertEqual(row2[4], 9.0)
        # Verify color styles (green for Technical Round)
        self.assertEqual(ws.cell(row=2, column=6).fill.start_color.rgb, '00FF00')




